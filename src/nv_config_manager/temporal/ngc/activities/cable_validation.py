# SPDX-FileCopyrightText: Copyright (c) 2024-2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: Apache-2.0
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# pylint: disable=too-many-locals.too-many-nested-blocks,too-many-branches
"""Activities for cable validation."""

from __future__ import annotations

import base64
import copy
import csv
import hashlib
import io
import re
from collections.abc import Iterable
from typing import Any

import netaddr
import pandas as pd
from openpyxl.utils import get_column_letter
from py_markdown_table.markdown_table import markdown_table
from pydantic import BaseModel, ConfigDict, Field
from temporalio import activity
from temporalio.exceptions import ApplicationError

from nv_config_manager.common.log import LogCategory, get_logger
from nv_config_manager.temporal.client.device import (
    DeviceArpTable,
    DeviceMacTable,
    DeviceNeighborData,
    InterfaceNeighborData,
    format_mac,
    is_mac_address,
)
from nv_config_manager.temporal.client.nautobot import NautobotClient
from nv_config_manager.temporal.common.mixins.device import NetworkDeviceData

logger = get_logger(__name__, category=LogCategory.TEMPORAL_ACTIVITY)

# Columns to exclude from CSV and markdown table output (alias/display names)
CSV_EXCLUDE_COLUMNS: frozenset[str] = frozenset({"Troubleshooting Info"})
MARKDOWN_EXCLUDE_COLUMNS: frozenset[str] = frozenset(
    {"Troubleshooting Info", "Start Rack", "Intended End Rack", "ID"}
)

# Issue message constants
LINK_UP_NO_NEIGHBOR_MSG = "Link is up but no neighbor found"
LINK_DOWN_MSG = "Link is down."
UNEXPECTED_CONNECTION_MSG = "Unexpected connection found"
INCORRECT_CABLING_PREFIX = "Incorrect cabling"

# Host summary tab (Excel) column headers, in display order. "Missing Cables" is
# the primary triage signal NVIS sorts on.
HOST_SUMMARY_COLUMNS: tuple[str, ...] = (
    "Host",
    "Rack",
    "Missing Cables",
    "Miscabled",
    "Unexpected Connections",
    "Total Issues",
)

# Excel sheet names for the site cable validation workbook.
DETAIL_SHEET_NAME = "Cable Issues"
HOST_SUMMARY_SHEET_NAME = "Host Summary"

EXCEL_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _xlsx_download_filename(site: str) -> str:
    """Build a friendly download filename like ``site-cable-validation-<site>.xlsx``.

    The site is slugified (lowercased, non-alphanumeric runs collapsed to ``-``)
    so the name is safe across browsers and filesystems. ``site`` is always
    present -- it is a required, non-empty field on the workflow input.
    """
    slug = re.sub(r"[^a-z0-9]+", "-", site.lower()).strip("-")
    return f"site-cable-validation-{slug}.xlsx"


def _rack_position_str(rack: str | None, position: int | None) -> str | None:
    """Return rack position string (e.g. 'rack1:u42') or None if rack/position missing."""
    if rack and position is not None:
        return f"{rack}:u{position}"
    return None


def _display_or_macs(primary: str | None, macs: list[str] | None) -> str | None:
    """Return primary if set, else comma-joined macs or None (for Actual End Device/Port)."""
    return primary or (",".join(macs) if macs else None)


class ValidateDeviceNeighborsInput(BaseModel):
    """Input for the device neighbor validation activity."""

    device: NetworkDeviceData
    intended: DeviceNeighborData
    actual: DeviceNeighborData
    mac_table: DeviceMacTable
    arp_table: DeviceArpTable


class InvalidCable(BaseModel, validate_assignment=True):
    """Invalid Cable for Validation Results."""

    intended: InterfaceNeighborData | None = None
    actual: InterfaceNeighborData | None = None


class ValidateDeviceNeighborsResult(BaseModel, validate_assignment=True):
    """Result for a Device Cable Validation."""

    # key is the interface name
    interfaces: dict[str, InvalidCable] = {}


class CableValidationRow(BaseModel):
    """A single row of cable validation results with serialization for CSV/markdown."""

    model_config = ConfigDict(populate_by_name=True)

    start_device: str | None = Field(default=None, alias="Start Device")
    start_port: str | None = Field(default=None, alias="Start Port")
    start_rack: str | None = Field(default=None, alias="Start Rack")
    intended_end_device: str | None = Field(default=None, alias="Intended End Device")
    intended_end_port: str | None = Field(default=None, alias="Intended End Port")
    intended_end_rack: str | None = Field(default=None, alias="Intended End Rack")
    actual_end_device: str | None = Field(default=None, alias="Actual End Device")
    actual_end_port: str | None = Field(default=None, alias="Actual End Port")
    issue: str | None = Field(default=None, alias="Issue")
    troubleshooting_info: str | None = Field(default=None, alias="Troubleshooting Info")
    id_: str | None = Field(default=None, alias="ID")

    def to_markdown(self) -> dict[str, Any]:
        """Return dict of columns to include in markdown table (excludes MARKDOWN_EXCLUDE_COLUMNS)."""
        return {
            k: v
            for k, v in self.model_dump(by_alias=True).items()
            if k not in MARKDOWN_EXCLUDE_COLUMNS
        }

    def to_csv_dict(self) -> dict[str, Any]:
        """Return dict of columns to include in CSV export (excludes CSV_EXCLUDE_COLUMNS)."""
        return {
            k: v for k, v in self.model_dump(by_alias=True).items() if k not in CSV_EXCLUDE_COLUMNS
        }

    @classmethod
    def compute_id(cls, row: CableValidationRow) -> str:
        """Compute hash ID for this row."""
        return hashlib.md5(
            (
                str(row.start_device)
                + str(row.start_port)
                + str(row.start_rack)
                + str(row.intended_end_device)
                + str(row.intended_end_port)
                + str(row.intended_end_rack)
                + str(row.actual_end_device)
                + str(row.actual_end_port)
            ).encode("utf-8"),
            usedforsecurity=False,
        ).hexdigest()


def _as_mac_address(s: str) -> str | None:
    try:
        return str(netaddr.EUI(s))
    except netaddr.core.AddrFormatError:
        return None


def _mac_matches_with_offset(actual_mac: str, expected_mac: str) -> bool:
    """
    Check if actual MAC matches expected MAC or expected MAC + 0x10 offset.

    DPUs sometimes advertise their MAC over LLDP with an offset of 0x10 when in NIC mode.
    This function checks both the original expected MAC and the offset version.

    Args:
        actual_mac: The MAC address received via LLDP
        expected_mac: The intended/expected MAC address

    Returns:
        True if actual_mac matches expected_mac or expected_mac + 0x10
    """
    try:
        actual_eui = netaddr.EUI(actual_mac)
        expected_eui = netaddr.EUI(expected_mac)

        # Check exact match first
        if actual_eui == expected_eui:
            return True

        # Check if actual MAC matches expected MAC + 0x10 offset
        offset_eui = netaddr.EUI(expected_eui.value + 0x10)
        return bool(actual_eui == offset_eui)

    except (netaddr.core.AddrFormatError, ValueError):
        return False


def _validate_neighbor_has_required_fields(
    intended_neighbor: InterfaceNeighborData, device: NetworkDeviceData
) -> None:
    """Validate that intended neighbor has required role and name fields."""
    if not intended_neighbor.device_role:
        raise ApplicationError(
            f"No role information for intended neighbor {intended_neighbor} on device {device}"
        )
    if not intended_neighbor.device_name:
        raise ApplicationError(
            f"No name information for intended neighbor {intended_neighbor} on device {device}"
        )


def _check_link_state(
    intended_interface: str,
    actual: DeviceNeighborData,
    intended_neighbor: InterfaceNeighborData,
) -> InvalidCable | None:
    """Check link state and return InvalidCable if link is down."""
    if not actual.link_states.get(intended_interface, False):
        return InvalidCable(
            intended=intended_neighbor,
            actual=InterfaceNeighborData(
                link_up=False,
                ts_info=actual.ts_info.get(intended_interface),
            ),
        )
    return None


def _check_interface_mac_match(
    actual_neighbor_interface: str | None, expected_mac: str | None
) -> bool:
    """Check if actual neighbor interface MAC matches expected MAC."""
    if not actual_neighbor_interface or not expected_mac:
        return False

    neighbor_if_mac = _as_mac_address(actual_neighbor_interface)
    if neighbor_if_mac:
        return _mac_matches_with_offset(neighbor_if_mac, expected_mac)
    return False


def _check_device_mac_match(actual_neighbor_device: str | None, expected_mac: str | None) -> bool:
    """Check if actual neighbor device MAC matches expected MAC."""
    if not actual_neighbor_device or not expected_mac:
        return False

    neighbor_device_mac = _as_mac_address(actual_neighbor_device)
    if neighbor_device_mac:
        return _mac_matches_with_offset(neighbor_device_mac, expected_mac)
    return False


def _check_standard_lldp_match(
    intended_neighbor_interface: str | None,
    intended_neighbor_device: str | None,
    actual_neighbor_interface: str | None,
    actual_neighbor_device: str | None,
) -> bool:
    """Check if LLDP data matches using standard interface and device names."""
    return (
        intended_neighbor_interface == actual_neighbor_interface
        and intended_neighbor_device == actual_neighbor_device
    )


def _check_lldp_validity(
    intended_interface: str,
    intended_neighbor: InterfaceNeighborData,
    actual: DeviceNeighborData,
) -> bool:
    """Check if LLDP data is valid."""
    if intended_interface not in actual.neighbors:
        return False

    actual_neighbor = actual.neighbors[intended_interface]

    # Normalize strings to lowercase for comparison
    intended_neighbor_interface = intended_neighbor.name.lower() if intended_neighbor.name else None
    intended_neighbor_device = (
        intended_neighbor.device_name.lower() if intended_neighbor.device_name else None
    )
    actual_neighbor_interface = actual_neighbor.name.lower() if actual_neighbor.name else None
    actual_neighbor_device = (
        actual_neighbor.device_name.lower() if actual_neighbor.device_name else None
    )

    expected_mac = intended_neighbor.macs[0] if intended_neighbor.macs else None

    # Device sent the interface mac instead of name
    if _check_interface_mac_match(actual_neighbor_interface, expected_mac):
        return True

    # Device sent mac address instead of device name
    if _check_device_mac_match(actual_neighbor_device, expected_mac):
        return True

    # Standard LLDP check
    return _check_standard_lldp_match(
        intended_neighbor_interface,
        intended_neighbor_device,
        actual_neighbor_interface,
        actual_neighbor_device,
    )


def _check_mac_validity(
    intended_interface: str,
    intended_neighbor: InterfaceNeighborData,
    mac_table: DeviceMacTable,
    arp_table: DeviceArpTable,
) -> bool:
    """Check if MAC address is valid."""
    if not intended_neighbor.macs:
        return False

    mac = intended_neighbor.macs[0]
    return (mac in mac_table.by_mac and mac_table.by_mac[mac].interface == intended_interface) or (
        mac in arp_table.interface_to_mac.get(intended_interface, [])
    )


def _build_actual_neighbor_data(
    intended_interface: str,
    intended_neighbor: InterfaceNeighborData,
    actual: DeviceNeighborData,
    mac_table: DeviceMacTable,
    arp_table: DeviceArpTable,
) -> InterfaceNeighborData:
    """Build actual neighbor data for error reporting."""
    mac_neighbor = InterfaceNeighborData(
        macs=(
            mac_table.by_interface[intended_interface]
            if intended_interface in mac_table.by_interface
            and mac_table.by_interface[intended_interface]
            else arp_table.interface_to_mac.get(intended_interface, [])
        ),
        link_up=actual.link_states.get(intended_interface, False),
    )

    # Prefer MAC error message if MAC data present
    if intended_neighbor.macs and (
        mac_table.by_interface.get(intended_interface)
        or arp_table.interface_to_mac.get(intended_interface)
    ):
        return mac_neighbor

    if intended_interface in actual.neighbors:
        return actual.neighbors[intended_interface]

    return mac_neighbor


def _process_intended_interface(
    intended_interface: str,
    intended_neighbor: InterfaceNeighborData,
    intended: DeviceNeighborData,
    actual: DeviceNeighborData,
    device: NetworkDeviceData,
    mac_table: DeviceMacTable,
    arp_table: DeviceArpTable,
) -> InvalidCable | None:
    """Process a single intended interface and return InvalidCable if invalid."""
    if intended_interface in intended.ignore:
        return None

    _validate_neighbor_has_required_fields(intended_neighbor, device)

    # Check link state
    link_state_result = _check_link_state(intended_interface, actual, intended_neighbor)
    if link_state_result:
        return link_state_result

    # Skip further checks if link state only validation is requested
    if intended_interface in intended.link_state_only:
        return None

    # Check LLDP and MAC validity
    lldp_valid = _check_lldp_validity(intended_interface, intended_neighbor, actual)
    mac_valid = _check_mac_validity(intended_interface, intended_neighbor, mac_table, arp_table)

    # If either is valid, the cable is valid
    if mac_valid or lldp_valid:
        return None

    # Build error data
    actual_neighbor = _build_actual_neighbor_data(
        intended_interface, intended_neighbor, actual, mac_table, arp_table
    )

    return InvalidCable(
        intended=intended_neighbor,
        actual=actual_neighbor,
    )


def _find_unexpected_neighbors(
    intended: DeviceNeighborData,
    actual: DeviceNeighborData,
) -> dict[str, InvalidCable]:
    """Find neighbors in actual that aren't in intended."""
    unexpected = {}

    for actual_interface, actual_neighbor in actual.neighbors.items():
        if actual_interface in intended.ignore:
            continue

        if (
            (actual_neighbor.device_name or actual_neighbor.macs or actual_neighbor.device_serial)
            and actual_interface not in intended.neighbors
            and actual_interface
        ):
            unexpected[actual_interface] = InvalidCable(actual=actual_neighbor)

    return unexpected


@activity.defn
async def validate_device_neighbors(
    activity_input: ValidateDeviceNeighborsInput,
) -> ValidateDeviceNeighborsResult:
    """
    Validate device neighbors.

    MAC and LLDP will be checked, if either is correct, the link is considered valid.

    This function will catch actual LLDP neighbors that aren't in nautobot, but it will
    not validate MAC addresses that are on the device but not in nautobot.

    """
    intended = activity_input.intended
    actual = activity_input.actual
    device = activity_input.device
    mac_table = activity_input.mac_table
    arp_table = activity_input.arp_table
    result = ValidateDeviceNeighborsResult()

    for intended_interface, intended_neighbor in intended.neighbors.items():
        invalid_cable = _process_intended_interface(
            intended_interface,
            intended_neighbor,
            intended,
            actual,
            device,
            mac_table,
            arp_table,
        )
        if invalid_cable:
            result.interfaces[intended_interface] = invalid_cable

    # Find unexpected neighbors (in actual but not in intended)
    unexpected = _find_unexpected_neighbors(intended, actual)
    result.interfaces.update(unexpected)

    return result


class CableValidationResultData(BaseModel):
    """Cable validation result data."""

    interfaces: dict[str, InvalidCable]
    device: NetworkDeviceData | None = None


class DecorateResultActivityInput(BaseModel):
    """Decorate result activity input."""

    devices: dict[str, CableValidationResultData]


class DecorateResultActivityOutput(BaseModel):
    """Decorate result activity input."""

    devices: dict[str, CableValidationResultData]


@activity.defn
async def decorate_result(
    activity_input: DecorateResultActivityInput,
) -> DecorateResultActivityOutput:
    """Decorate result activity."""
    activity_result = copy.deepcopy(activity_input.devices)
    mac_to_host: dict[str, tuple[str, str | None]] = {}
    for device_result in activity_result.values():
        for interface_result in device_result.interfaces.values():
            macs = interface_result.actual.macs if interface_result.actual else []
            for mac in macs:
                if mac not in mac_to_host:
                    mac_to_host[mac] = (mac, None)
            if (
                interface_result.actual
                and interface_result.actual.name
                and is_mac_address(interface_result.actual.name)
                and interface_result.actual.name not in mac_to_host
            ):
                mac_to_host[interface_result.actual.name] = (
                    interface_result.actual.name,
                    None,
                )

    if not mac_to_host:
        return DecorateResultActivityOutput(devices=activity_result)

    client = NautobotClient()
    async with client:
        interfaces = await client.get_interfaces_by_mac(mac_addresses=list(mac_to_host.keys()))

    for interface in interfaces:
        if interface.mac_address:
            mac_to_host[interface.mac_address] = (interface.name, interface.host)

    for device_result in activity_result.values():
        for interface_result in device_result.interfaces.values():
            if interface_result.actual:
                for mac in interface_result.actual.macs:
                    if mac in mac_to_host:
                        if not interface_result.actual.name:
                            interface_result.actual.name = mac_to_host[mac][0]
                        if not interface_result.actual.device_name:
                            interface_result.actual.device_name = mac_to_host[mac][1]
                        # Assume one connected host match per interface
                        break
            if (
                interface_result.actual
                and interface_result.actual.name
                and is_mac_address(interface_result.actual.name)
                and interface_result.actual.name in mac_to_host
            ):
                interface_result.actual.device_name = mac_to_host[interface_result.actual.name][1]
                interface_result.actual.name = mac_to_host[interface_result.actual.name][0]

    return DecorateResultActivityOutput(devices=activity_result)


def _row_unexpected_connection(
    device_name: str,
    interface_name: str,
    device_rack_pos: str | None,
    actual: InterfaceNeighborData,
    actual_device_name: str | None,
) -> CableValidationRow:
    """Build row for unexpected connection (actual neighbor not in intended)."""
    return CableValidationRow(  # type: ignore[call-arg]
        start_device=device_name,
        start_port=interface_name,
        start_rack=device_rack_pos,
        intended_end_device=None,
        intended_end_port=None,
        intended_end_rack=None,
        actual_end_device=_display_or_macs(actual_device_name, actual.macs),
        actual_end_port=_display_or_macs(actual.name, actual.macs),
        issue="Unexpected connection found",
        troubleshooting_info=actual.ts_info,
    )


def _row_link_down(
    device_name: str,
    interface_name: str,
    device_rack_pos: str | None,
    intended_device_name: str | None,
    intended: InterfaceNeighborData,
    actual: InterfaceNeighborData,
) -> CableValidationRow:
    """Build row for link down."""
    return CableValidationRow(  # type: ignore[call-arg]
        start_device=device_name,
        start_port=interface_name,
        start_rack=device_rack_pos,
        intended_end_device=intended_device_name,
        intended_end_port=intended.name,
        intended_end_rack=_rack_position_str(intended.device_rack, intended.device_position),
        actual_end_device=None,
        actual_end_port=None,
        issue="Link is down.",
        troubleshooting_info=actual.ts_info,
    )


def _row_mac_mismatch(
    device_name: str,
    interface_name: str,
    device_rack_pos: str | None,
    intended_device_name: str | None,
    intended: InterfaceNeighborData,
    actual: InterfaceNeighborData,
    actual_device_name: str | None,
) -> CableValidationRow:
    """Build row for incorrect cabling based on MAC comparison."""
    return CableValidationRow(  # type: ignore[call-arg]
        start_device=device_name,
        start_port=interface_name,
        start_rack=device_rack_pos,
        intended_end_device=intended_device_name,
        intended_end_port=intended.name,
        intended_end_rack=_rack_position_str(intended.device_rack, intended.device_position),
        actual_end_device=_display_or_macs(actual_device_name, actual.macs),
        actual_end_port=_display_or_macs(actual.name, actual.macs),
        issue=(
            "Incorrect cabling, actual should match intended. "
            f"Based on expected MAC {intended.macs[0]}*"
        ),
        troubleshooting_info=actual.ts_info,
    )


def _row_lldp_mismatch(
    device_name: str,
    interface_name: str,
    device_rack_pos: str | None,
    intended_device_name: str | None,
    intended: InterfaceNeighborData,
    actual: InterfaceNeighborData,
    actual_device_name: str | None,
) -> CableValidationRow:
    """Build row for incorrect cabling based on LLDP (actual.name set)."""
    end_port = (
        f"{intended.name} ({format_mac(intended.macs[0])})"
        if is_mac_address(actual.name) and intended.macs
        else intended.name
    )
    return CableValidationRow(  # type: ignore[call-arg]
        start_device=device_name,
        start_port=interface_name,
        start_rack=device_rack_pos,
        intended_end_device=intended_device_name,
        intended_end_port=end_port,
        intended_end_rack=_rack_position_str(intended.device_rack, intended.device_position),
        actual_end_device=_display_or_macs(actual_device_name, actual.macs),
        actual_end_port=_display_or_macs(actual.name, actual.macs),
        issue=("Incorrect cabling, actual should match intended. Based on LLDP data"),
        troubleshooting_info=actual.ts_info,
    )


def _row_link_up_no_neighbor(
    device_name: str,
    interface_name: str,
    device_rack_pos: str | None,
    intended_device_name: str | None,
    intended: InterfaceNeighborData,
    actual: InterfaceNeighborData,
) -> CableValidationRow:
    """Build row for link up but no neighbor found."""
    return CableValidationRow(  # type: ignore[call-arg]
        start_device=device_name,
        start_port=interface_name,
        start_rack=device_rack_pos,
        intended_end_device=intended_device_name,
        intended_end_port=intended.name,
        intended_end_rack=_rack_position_str(intended.device_rack, intended.device_position),
        actual_end_device=_display_or_macs(None, actual.macs),
        actual_end_port=_display_or_macs(None, actual.macs),
        issue=LINK_UP_NO_NEIGHBOR_MSG,
        troubleshooting_info=actual.ts_info,
    )


def _format_single_interface_row(
    device_name: str,
    device_rack_pos: str | None,
    interface_name: str,
    interface_result: InvalidCable,
) -> CableValidationRow | None:
    """Convert a single interface validation result into a table row.

    Args:
        device_name: Lowercase device name
        device_rack_pos: Device rack position string (e.g., "rack1:u42")
        interface_name: Interface name
        interface_result: InvalidCable result for this interface

    Returns:
        CableValidationRow with ID set, or None if this should be skipped
    """
    intended = interface_result.intended
    actual = interface_result.actual
    if not actual:
        return None

    intended_device_name = (
        intended.device_name.lower() if intended and intended.device_name else None
    )
    actual_device_name = actual.device_name.lower() if actual.device_name else None

    if not intended:
        row = _row_unexpected_connection(
            device_name, interface_name, device_rack_pos, actual, actual_device_name
        )
    elif actual.link_up is False:
        row = _row_link_down(
            device_name, interface_name, device_rack_pos, intended_device_name, intended, actual
        )
    elif intended.macs and actual.macs:
        row = _row_mac_mismatch(
            device_name,
            interface_name,
            device_rack_pos,
            intended_device_name,
            intended,
            actual,
            actual_device_name,
        )
    elif actual.name:
        row = _row_lldp_mismatch(
            device_name,
            interface_name,
            device_rack_pos,
            intended_device_name,
            intended,
            actual,
            actual_device_name,
        )
    else:
        row = _row_link_up_no_neighbor(
            device_name, interface_name, device_rack_pos, intended_device_name, intended, actual
        )

    row.id_ = CableValidationRow.compute_id(row)
    return row


def _should_skip_link_down_dedup(
    row: CableValidationRow, dedup: set[tuple[str, str]] | None
) -> bool:
    """Return True if row should be skipped (duplicate link-down). Updates dedup."""
    if dedup is None:
        return False
    if row.issue != "Link is down.":
        return False
    intended_device = row.intended_end_device
    intended_port = row.intended_end_port
    if not intended_device or not intended_port:
        return False
    intended_end = (intended_device, intended_port)
    if intended_end in dedup:
        return True
    dedup.add(intended_end)
    return False


def _should_skip_no_neighbor(row: CableValidationRow, ignore_no_neighbor: bool) -> bool:
    """Return True if row should be skipped (link up, no neighbor)."""
    return bool(ignore_no_neighbor and row.issue == LINK_UP_NO_NEIGHBOR_MSG)


def _format_device_result_row(
    device: NetworkDeviceData | None,
    interfaces: dict[str, InvalidCable],
    dedup: set[tuple[str, str]] | None = None,
    ignore_no_neighbor: bool = False,
) -> list[CableValidationRow]:
    """Convert device interface validation results into table rows.

    Args:
        device: NetworkDeviceData object (can be None)
        interfaces: Dictionary of interface names to InvalidCable results
        dedup: Optional set for deduplicating "Link is down" cases
        ignore_no_neighbor: If True, skip "Link is up but no neighbor found" cases

    Returns:
        List of CableValidationRow with IDs set
    """
    device_name = device.name.lower() if device and device.name else "unknown"
    device_rack_pos = _rack_position_str(device.rack, device.position) if device else None

    results: list[CableValidationRow] = []
    for interface_name, interface_result in interfaces.items():
        row = _format_single_interface_row(
            device_name, device_rack_pos, interface_name, interface_result
        )
        if row is None:
            continue
        if _should_skip_link_down_dedup(row, dedup):
            continue
        if _should_skip_no_neighbor(row, ignore_no_neighbor):
            continue
        results.append(row)
    return results


def _generate_notes(results: list[CableValidationRow]) -> list[str]:
    """Generate notes based on patterns in validation results.

    Args:
        results: List of CableValidationRow

    Returns:
        List of note strings
    """
    notes = []
    for row in results:
        issue = row.issue or ""
        if "*" in issue:
            notes.append(
                "*Either the expected MAC in our database is wrong"
                ", or this link is not cabled correctly."
            )
            break
    return notes


def _format_results_markdown(
    results: list[CableValidationRow],
    csv_exclude_columns: Iterable[str],
    markdown_exclude_columns: Iterable[str],
    empty_message: str = "No invalid cabling found.",
    max_display_results: int = 1000,
    export: str = "csv",
    summary_devices: dict[str, CableValidationResultData] | None = None,
    export_filename: str | None = None,
) -> str:
    """Format validation results into markdown with an export link, table, and notes.

    Args:
        results: List of CableValidationRow
        csv_exclude_columns: Columns to exclude from CSV export (unused; model uses CSV_EXCLUDE_COLUMNS)
        markdown_exclude_columns: Columns to exclude from markdown table (unused; model uses MARKDOWN_EXCLUDE_COLUMNS)
        empty_message: Message to display when there are no results
        max_display_results: Maximum number of results to display in table
        export: "csv" for a single CSV link, "xlsx" for a two-tab Excel download
        summary_devices: All queried switches, used to seed the Excel summary tab so
            healthy switches appear with zero counts
        export_filename: Friendly download filename for the Excel link

    Returns:
        Formatted markdown string
    """
    if not results:
        return empty_message

    markdown_rows = [r.to_markdown() for r in results]
    export_link = (
        _generate_xlsx_link(results, summary_devices, export_filename)
        if export == "xlsx"
        else _generate_csv_link(results)
    )

    markdown = f"{export_link}\n"

    if len(results) > max_display_results:
        markdown += (
            f"Too many results to display ({len(results)} errors), please export to CSV to view.\n"
        )
    else:
        markdown += str(
            markdown_table(markdown_rows).set_params(quote=False, row_sep="markdown").get_markdown()
        )

    notes = _generate_notes(results)
    if notes:
        markdown += "\n\n" + "\n".join(notes)

    return markdown


def _classify_issue(issue: str | None) -> str:
    """Bucket a row's issue into a host-summary category.

    Returns one of: "missing", "miscabled", "unexpected", "other". "missing"
    covers links that should exist but don't.
    """
    if issue in (LINK_DOWN_MSG, LINK_UP_NO_NEIGHBOR_MSG):
        return "missing"
    if issue and issue.startswith(INCORRECT_CABLING_PREFIX):
        return "miscabled"
    if issue == UNEXPECTED_CONNECTION_MSG:
        return "unexpected"
    return "other"


def _build_host_summary(
    results: list[CableValidationRow],
    devices: dict[str, CableValidationResultData] | None = None,
) -> list[dict[str, Any]]:
    """Aggregate per-switch issue counts for the summary tab."""
    summary: dict[str, dict[str, Any]] = {}

    def _get_bucket(host: str, rack: str | None) -> dict[str, Any]:
        bucket = summary.get(host)
        if bucket is None:
            bucket = {
                "Host": host,
                "Rack": rack,
                "Missing Cables": 0,
                "Miscabled": 0,
                "Unexpected Connections": 0,
                "Total Issues": 0,
            }
            summary[host] = bucket
        elif not bucket["Rack"] and rack:
            bucket["Rack"] = rack
        return bucket

    # Seed a row for every queried switch so healthy ones still appear (0 counts).
    # Match the host key to how rows are keyed (_format_device_result_row lowercases
    # the device name) so a switch is never listed twice.
    for name, data in (devices or {}).items():
        device = data.device
        host = device.name.lower() if device and device.name else name
        rack = _rack_position_str(device.rack, device.position) if device else None
        _get_bucket(host, rack)

    for row in results:
        bucket = _get_bucket(row.start_device or "(unknown)", row.start_rack)
        bucket["Total Issues"] += 1
        category = _classify_issue(row.issue)
        if category == "missing":
            bucket["Missing Cables"] += 1
        elif category == "miscabled":
            bucket["Miscabled"] += 1
        elif category == "unexpected":
            bucket["Unexpected Connections"] += 1

    return sorted(
        summary.values(),
        key=lambda r: (-r["Missing Cables"], -r["Total Issues"], r["Host"]),
    )


def _style_excel_sheet(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    """Apply uniform column widths and an autofilter to a worksheet."""
    worksheet = writer.sheets[sheet_name]
    for col_idx in range(len(df.columns)):
        worksheet.column_dimensions[get_column_letter(col_idx + 1)].width = 22
    if len(df) > 0:
        last_col = get_column_letter(len(df.columns))
        worksheet.auto_filter.ref = f"A1:{last_col}{len(df) + 1}"


def _build_cable_validation_workbook(
    results: list[CableValidationRow],
    devices: dict[str, CableValidationResultData] | None = None,
) -> bytes:
    """Build a two-tab .xlsx: full cable issue detail plus a per-switch summary."""
    detail_rows = [r.to_csv_dict() for r in results]
    detail_columns = (
        list(detail_rows[0].keys())
        if detail_rows
        else list(CableValidationRow().to_csv_dict().keys())  # type: ignore[call-arg]
    )
    detail_df = pd.DataFrame(detail_rows, columns=detail_columns)
    summary_df = pd.DataFrame(
        _build_host_summary(results, devices), columns=list(HOST_SUMMARY_COLUMNS)
    )

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        detail_df.to_excel(writer, sheet_name=DETAIL_SHEET_NAME, index=False)
        _style_excel_sheet(writer, DETAIL_SHEET_NAME, detail_df)
        summary_df.to_excel(writer, sheet_name=HOST_SUMMARY_SHEET_NAME, index=False)
        _style_excel_sheet(writer, HOST_SUMMARY_SHEET_NAME, summary_df)
    return buffer.getvalue()


def _generate_xlsx_link(
    results: list[CableValidationRow],
    devices: dict[str, CableValidationResultData] | None = None,
    filename: str | None = None,
) -> str:
    """Generate an Excel download link with detail and host-summary tabs."""
    workbook = _build_cable_validation_workbook(results, devices)
    b64_xlsx = base64.b64encode(workbook).decode("utf-8")
    fragment = f"#filename={filename}" if filename else ""
    return f"[Download Excel](data:{EXCEL_MIME_TYPE};base64,{b64_xlsx}{fragment})"


def _generate_csv_link(results: list[CableValidationRow]) -> str:
    """Generate a CSV export link from validation results.

    Args:
        results: List of CableValidationRow

    Returns:
        CSV data URI string for download (markdown link)
    """
    if not results:
        return "[Export to CSV](data:text/csv;base64,)"

    csv_rows = [r.to_csv_dict() for r in results]
    result_csv = io.StringIO()
    writer = csv.DictWriter(result_csv, csv_rows[0].keys())
    writer.writeheader()
    writer.writerows(csv_rows)
    csv_string = result_csv.getvalue()
    b64_csv = base64.b64encode(csv_string.encode("utf-8"))
    return f"[Export to CSV](data:text/csv;base64,{b64_csv.decode()})"


class FormatResultsActivityInput(BaseModel):
    """Format results activity input."""

    devices: dict[str, CableValidationResultData]
    failed_devices: dict[str, str]
    ignore_no_neighbor: bool = False
    site: str


@activity.defn
def format_results(activity_input: FormatResultsActivityInput) -> str:
    """Format Cable Validation Results in markdown."""
    results = []
    dedup: set[tuple[str, str]] = set()
    for _, device_results in activity_input.devices.items():
        if device_results.interfaces:
            device_rows = _format_device_result_row(
                device_results.device,
                device_results.interfaces,
                dedup=dedup,
                ignore_no_neighbor=activity_input.ignore_no_neighbor,
            )
            results.extend(device_rows)

    markdown = ""
    if activity_input.failed_devices:
        failures = []
        markdown += "### Failed Devices\n"
        markdown += "Address the listed issues and re-run the workflow for complete results.\n\n"
        for device_name, error in activity_input.failed_devices.items():
            failures.append(
                {
                    "Failed Device": device_name,
                    "Reason": error,
                }
            )
        markdown += (
            markdown_table(failures).set_params(quote=False, row_sep="markdown").get_markdown()
        )
        markdown += "\n\n"

    markdown += _format_results_markdown(
        results,
        csv_exclude_columns=CSV_EXCLUDE_COLUMNS,
        markdown_exclude_columns=MARKDOWN_EXCLUDE_COLUMNS,
        export="xlsx",
        summary_devices=activity_input.devices,
        export_filename=_xlsx_download_filename(activity_input.site),
    )
    return markdown


class FormatDeviceValidationResultInput(BaseModel):
    """Format device validation result activity input."""

    device: NetworkDeviceData
    validation_result: ValidateDeviceNeighborsResult
    ignore_no_neighbor: bool = False


@activity.defn
def format_device_validation_result(
    activity_input: FormatDeviceValidationResultInput,
) -> str:
    """Format a single device's cable validation results in markdown."""
    results = _format_device_result_row(
        activity_input.device,
        activity_input.validation_result.interfaces,
        ignore_no_neighbor=activity_input.ignore_no_neighbor,
    )

    return _format_results_markdown(
        results,
        csv_exclude_columns=CSV_EXCLUDE_COLUMNS,
        markdown_exclude_columns=MARKDOWN_EXCLUDE_COLUMNS,
        empty_message="All cable connections are valid.",
        max_display_results=1000,
    )
