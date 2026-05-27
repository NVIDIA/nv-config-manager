# SPDX-FileCopyrightText: Copyright (c) 2024-2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: Apache-2.0
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""Hardware validation activities."""

import base64
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from temporalio import activity

from nv_config_manager.temporal.client.device import NetworkConnection
from nv_config_manager.temporal.common.mixins.device import NetworkDeviceData


class HardwareValidationInput(BaseModel):
    """Hardware validation activity input."""

    device_data: NetworkDeviceData


class HardwareValidationOutput(BaseModel):
    """Hardware validation activity output for all API calls."""

    info: dict


class CompleteHardwareValidationOutput(BaseModel):
    """Complete hardware validation output containing all collected data."""

    device: NetworkDeviceData
    platform: dict
    fan: dict
    led: dict
    psu: dict
    voltage: dict
    inventory: dict


class HardwareValidationResult(BaseModel):
    """Simple hardware validation result."""

    success: bool
    devices_validated: int
    total_entries: int
    message: str


class CreateExcelInput(BaseModel):
    """Input for Excel generation activity."""

    command_name: str
    devices_data_and_results: dict[str, dict[str, Any]]


class CreateExcelOutput(BaseModel):
    """Output for Excel generation activity."""

    excel_data: str
    row_count: int


class CreateConsolidatedExcelInput(BaseModel):
    """Input for consolidated Excel generation with multiple worksheets."""

    stage_data: dict[str, dict[str, dict[str, Any]]]


class CreateConsolidatedExcelOutput(BaseModel):
    """Output for consolidated Excel generation."""

    excel_data: str
    total_row_count: int
    worksheet_counts: dict[str, int]


@activity.defn
def get_platform(
    activity_input: HardwareValidationInput,
) -> HardwareValidationOutput:
    """Get platform information from the device."""
    connection = NetworkConnection.from_device_data(activity_input.device_data)
    platform_info = connection.get_platform()
    return HardwareValidationOutput(info=platform_info)


@activity.defn
def get_platform_environment_fan(
    activity_input: HardwareValidationInput,
) -> HardwareValidationOutput:
    """Get platform fan information from the device."""
    connection = NetworkConnection.from_device_data(activity_input.device_data)
    fan_info = connection.get_platform_environment_fan()
    return HardwareValidationOutput(info=fan_info)


@activity.defn
def get_platform_environment_led(
    activity_input: HardwareValidationInput,
) -> HardwareValidationOutput:
    """Get platform LED information from the device."""
    connection = NetworkConnection.from_device_data(activity_input.device_data)
    led_info = connection.get_platform_environment_led()
    return HardwareValidationOutput(info=led_info)


@activity.defn
def get_platform_environment_psu(
    activity_input: HardwareValidationInput,
) -> HardwareValidationOutput:
    """Get platform PSU information from the device."""
    connection = NetworkConnection.from_device_data(activity_input.device_data)
    psu_info = connection.get_platform_environment_psu()
    return HardwareValidationOutput(info=psu_info)


@activity.defn
def get_platform_environment_voltage(
    activity_input: HardwareValidationInput,
) -> HardwareValidationOutput:
    """Get platform voltage information from the device."""
    connection = NetworkConnection.from_device_data(activity_input.device_data)
    voltage_info = connection.get_platform_environment_voltage()
    return HardwareValidationOutput(info=voltage_info)


@activity.defn
def get_platform_inventory(
    activity_input: HardwareValidationInput,
) -> HardwareValidationOutput:
    """Get platform inventory information from the device."""
    connection = NetworkConnection.from_device_data(activity_input.device_data)
    inventory_info = connection.get_platform_inventory()
    return HardwareValidationOutput(info=inventory_info)


@activity.defn
def create_excel_export(activity_input: CreateExcelInput) -> CreateExcelOutput:
    """Create Excel export for stage data."""

    command_name = activity_input.command_name
    devices_data_and_results = activity_input.devices_data_and_results

    rows = []

    for result in devices_data_and_results.values():
        device_data = result["device_data"]
        command_result = result.get("command_result", {})

        if result.get("error"):
            continue

        if isinstance(device_data, dict):
            device_name = device_data.get("name", "N/A")
            rack_name = device_data.get("rack", "N/A") or "N/A"
            rack_position = device_data.get("position", "N/A") or "N/A"
        else:
            device_name = device_data.name
            rack_name = device_data.rack or "N/A"
            rack_position = device_data.position or "N/A"

        if command_name == "platform":
            # For platform, each top-level key becomes a row
            for value in command_result.values():
                if isinstance(value, dict):
                    flattened_data = _flatten_dict(value)
                    row = {
                        "device_name": device_name,
                        "rack_name": rack_name,
                        "rack_position": str(rack_position),
                        **flattened_data,
                    }
                    rows.append(row)
        else:
            # For other commands (fan, led, psu, voltage, inventory), each item becomes a row
            for item_name, item_data in command_result.items():
                if isinstance(item_data, dict):
                    flattened_data = _flatten_dict(item_data)
                    row = {
                        "device_name": device_name,
                        "rack_name": rack_name,
                        "rack_position": str(rack_position),
                        "item_name": item_name,
                        **flattened_data,
                    }
                    rows.append(row)

    if rows:
        df = pd.DataFrame(rows)
        for col in df.columns:
            if df[col].dtype in ["int64", "int32", "float64", "float32"]:
                continue
            else:
                df[col] = df[col].fillna("N/A")

        priority_columns = ["device_name", "rack_name", "rack_position"]
        existing_priority = [col for col in priority_columns if col in df.columns]
        other_columns = sorted([col for col in df.columns if col not in priority_columns])
        ordered_columns = existing_priority + other_columns
        df = df[ordered_columns]

        # Format column headers for better readability
        formatted_columns = {col: _format_column_header(col) for col in df.columns}
        df = df.rename(columns=formatted_columns)

        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=command_name.capitalize(), index=False)

            worksheet = writer.sheets[command_name.capitalize()]

            for col_idx in range(len(df.columns)):
                col_letter = get_column_letter(col_idx + 1)
                worksheet.column_dimensions[col_letter].width = 20

        excel_buffer.seek(0)
        excel_data = excel_buffer.getvalue()
        row_count = len(df)
    else:
        df = pd.DataFrame(
            columns=["device_name", "rack_name", "rack_position", "No data available"]
        )
        # Format column headers for better readability
        formatted_columns = {col: _format_column_header(col) for col in df.columns}
        df = df.rename(columns=formatted_columns)

        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            df.to_excel(writer, index=False)

            worksheet = writer.sheets["Sheet1"]
            for col_idx in range(len(df.columns)):
                col_letter = get_column_letter(col_idx + 1)
                worksheet.column_dimensions[col_letter].width = 20

        excel_buffer.seek(0)
        excel_data = excel_buffer.getvalue()
        row_count = 0

    excel_data_b64 = base64.b64encode(excel_data).decode("utf-8")
    return CreateExcelOutput(excel_data=excel_data_b64, row_count=row_count)


@activity.defn
def create_consolidated_excel_export(
    activity_input: CreateConsolidatedExcelInput,
) -> CreateConsolidatedExcelOutput:
    """Create consolidated Excel export with a worksheet for each stage."""

    excel_buffer = BytesIO()
    worksheet_counts = {}
    total_row_count = 0

    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
        for stage_name, devices_data_and_results in activity_input.stage_data.items():
            rows = []

            for result in devices_data_and_results.values():
                device_data = result["device_data"]
                command_result = result.get("command_result", {})

                if result.get("error"):
                    continue

                if isinstance(device_data, dict):
                    device_name = device_data.get("name", "N/A")
                    rack_name = device_data.get("rack", "N/A") or "N/A"
                    rack_position = device_data.get("position", "N/A") or "N/A"
                else:
                    device_name = device_data.name
                    rack_name = device_data.rack or "N/A"
                    rack_position = device_data.position or "N/A"

                if stage_name == "platform":
                    if isinstance(command_result, dict):
                        flattened_data = _flatten_dict(command_result)
                        row = {
                            "device_name": device_name,
                            "rack_name": rack_name,
                            "rack_position": str(rack_position),
                            **flattened_data,
                        }
                        rows.append(row)
                else:
                    if isinstance(command_result, dict):
                        for item_name, item_data in command_result.items():
                            if isinstance(item_data, dict):
                                flattened_data = _flatten_dict(item_data)
                                row = {
                                    "device_name": device_name,
                                    "rack_name": rack_name,
                                    "rack_position": str(rack_position),
                                    "item_name": item_name,
                                    **flattened_data,
                                }
                                rows.append(row)

            if rows:
                df = pd.DataFrame(rows)
                for col in df.columns:
                    if df[col].dtype in ["int64", "int32", "float64", "float32"]:
                        continue
                    else:
                        df[col] = df[col].fillna("N/A")

                priority_columns = ["device_name", "rack_name", "rack_position"]
                if "item_name" in df.columns:
                    priority_columns.append("item_name")
                existing_priority = [col for col in priority_columns if col in df.columns]
                other_columns = sorted([col for col in df.columns if col not in priority_columns])
                ordered_columns = existing_priority + other_columns
                df = df[ordered_columns]

                # Sort by rack_name and rack_position for logical physical ordering
                sort_columns = []
                if "rack_name" in df.columns:
                    sort_columns.append("rack_name")
                if "rack_position" in df.columns:
                    sort_columns.append("rack_position")
                if sort_columns:
                    df = df.sort_values(sort_columns)

                # Format column headers for better readability
                formatted_columns = {col: _format_column_header(col) for col in df.columns}
                df = df.rename(columns=formatted_columns)

                sheet_name = stage_name.capitalize()
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                worksheet = writer.sheets[sheet_name]
                for col_idx in range(len(df.columns)):
                    col_letter = get_column_letter(col_idx + 1)
                    worksheet.column_dimensions[col_letter].width = 20

                # Add autofilter to all columns
                if len(df) > 0:
                    max_col = len(df.columns)
                    max_row = len(df) + 1  # +1 for header row
                    worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

                worksheet_counts[sheet_name] = len(df)
                total_row_count += len(df)
            else:
                df = pd.DataFrame(
                    columns=[
                        "device_name",
                        "rack_name",
                        "rack_position",
                        "No data available",
                    ]
                )
                # Format column headers for better readability
                formatted_columns = {col: _format_column_header(col) for col in df.columns}
                df = df.rename(columns=formatted_columns)

                sheet_name = stage_name.capitalize()
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                worksheet = writer.sheets[sheet_name]
                for col_idx in range(len(df.columns)):
                    col_letter = get_column_letter(col_idx + 1)
                    worksheet.column_dimensions[col_letter].width = 20

                # Add autofilter to empty worksheet headers
                if len(df.columns) > 0:
                    max_col = len(df.columns)
                    worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}1"

                worksheet_counts[sheet_name] = 0

    excel_buffer.seek(0)
    excel_data = excel_buffer.getvalue()
    excel_data_b64 = base64.b64encode(excel_data).decode("utf-8")

    return CreateConsolidatedExcelOutput(
        excel_data=excel_data_b64,
        total_row_count=total_row_count,
        worksheet_counts=worksheet_counts,
    )


def _format_column_header(header: str) -> str:
    """Format column header for better readability."""
    # Replace underscores and hyphens with spaces
    formatted = header.replace("_", " ").replace("-", " ")
    # Capitalize each word
    formatted = " ".join(word.capitalize() for word in formatted.split())
    return formatted


def _flatten_dict(data: dict[str, Any], parent_key: str = "", sep: str = "_") -> dict[str, Any]:
    """Flatten a nested dictionary."""
    items: list[tuple[str, Any]] = []
    for k, v in data.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(_flatten_dict(v, new_key, sep=sep).items())
        else:
            items.append((new_key, v))
    return dict(items)
